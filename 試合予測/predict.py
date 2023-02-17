
import openpyxl
import re
import time
import pandas as pd
import numpy as np
import matplotlib
import matplotlib.pyplot as plt
import seaborn as sns



from sklearn.model_selection import train_test_split
from sklearn.svm import LinearSVC
from sklearn.preprocessing import StandardScaler


def getdata ():
    wb = openpyxl.load_workbook(filename="ldata_3data.xlsx")

    sheet = wb.worksheets[0]
    
    when = []
    teamp1 = []
    teamp2 = []
    weather = []
    humid = []
    temp = []
    result = []



    for i in range(2,3391):

        w1 = sheet.cell(row=i,column=2).value
        w2 = sheet.cell(row=i,column=3).value
        w3 = sheet.cell(row=i,column=5).value
        w4 = sheet.cell(row=i,column=7).value
        w5 = sheet.cell(row=i,column=8).value
        w6 = sheet.cell(row=i,column=9).value
        r = sheet.cell(row=i,column=11).value
        
    
        when.append(w1)
        teamp1.append(w2)
        teamp2.append(w3)
        weather.append(w4)
        humid.append(w5)
        temp.append(w6)
        result.append(r)
    
    
    
    wb.close()
    
    return when, teamp1, teamp2, weather, humid, temp, result

def makedata(a,b,c,d,e,f):

    apd = pd.DataFrame(a,columns=["節"])
    bpd = pd.DataFrame(b,columns=["ホームチーム"])
    cpd = pd.DataFrame(c,columns=["アウェイチーム"])
    dpd = pd.DataFrame(d,columns=["天気"])
    epd = pd.DataFrame(e,columns=["湿度"])
    fpd = pd.DataFrame(f,columns=["気温"])

    all = pd.concat([apd,bpd,cpd,dpd,epd,fpd],axis=1)

    return all
    
def predict(pre):

    when, temp1, temp2, weather, humid, temp, result =getdata()

    ldata_pd = makedata(when,temp1,temp2,weather,humid,temp)    

    result_pd = pd.DataFrame(result)

    #以下実装

    x_train,x_test,y_train,y_test = train_test_split(ldata_pd,result_pd,test_size=0.3,random_state=0)

    model = LinearSVC()
    scaler = StandardScaler()

    scaler.fit(x_train)
    x_train_std = scaler.transform(x_train)
    x_test_std = scaler.transform(x_test)

    model.fit(x_train_std,y_train)


    ans = model.predict(pre)

    return ans
